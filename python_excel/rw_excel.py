#!/usr/bin/env python
# -*- coding:UTF-8 -*-
# @File: rw_excel.py
# @Time: 2019/09/30 22:55:59
# @Author: moli
# @Version: python3
# @Contact: guiczhang@163.com
# @blog: https://blog.csdn.net/u014421797

# DESCRIPTION: 读写xlsx,格式化日期时间输出

import datetime
import os,sys

import numpy as np
import xlrd
from numpy import linalg as la
from xlutils.copy import copy

import openpyxl

# windows下改变根路径到执行程序路径下
os.chdir(sys.path[0])

def compared_data(readfile, writefile):
    '''
    读取数据
    将两个表中的数据进行对比,将源文件时间写入目标文件
    '''
    # 读源文件
    book = xlrd.open_workbook(readfile)
    org_read_sheet = book.sheet_by_index(0)

    # 读目标文件
    book = xlrd.open_workbook(writefile)
    drc_read_sheet = book.sheet_by_index(0)

    # 要写入的目标文件
    workbook = openpyxl.load_workbook(writefile)
    write_sheet = workbook["Sheet1"]

    org_nrows = org_read_sheet.nrows    # 总行数
    drc_nrows = drc_read_sheet.nrows

    for i in range(org_nrows):  # 源文件循环
        # 需要对比的数据
        flag = 1
        org_row_value = org_read_sheet.row_values(i)
        drc_row_value = drc_read_sheet.row_values(i)

        while((org_row_value[1:] != drc_row_value[1:]) and flag == 1):
            i += 1
            if i < drc_nrows:
                drc_row_value = drc_read_sheet.row_values(i)
                flag = 1
            else:
                flag = 0
                break
        else:
            # 两行相等，将时间填入新表中
            # print(org_row_value[0])
            if (flag == 1):
                # 修改0行1列的数据为'Haha'
                write_sheet.cell(row=i+1, column=1).value = org_row_value[0]

    filename = "./f2_mid.xlsx"
    workbook.save(filename=filename)
    return filename


def add_time(filename):
    '''
    添加空白时间
    '''
    # 写入目标文件
    workbook = openpyxl.load_workbook(filename)
    write_sheet = workbook["Sheet1"]

    book = xlrd.open_workbook(filename)
    # 读数据
    read_sheet = book.sheet_by_index(0)

    rows = read_sheet.nrows  # 获取行数

    # 循环读取时间
    i = 1
    while(i < rows):
        pre_value = read_sheet.row_values(i-1)  # 前一段时间
        pre_time = pre_value[0]
        now_value = read_sheet.row_values(i)
        now_time = now_value[0]
        # 判断时间是否为空
        second_time = 1  # 增加的秒数
        while(pre_time != '' and now_time == ''):
            time0 = datetime.datetime.strptime(
                pre_time.split('.')[0], '%Y/%m/%d %H:%M:%S')
            time1 = (time0 + datetime.timedelta(seconds=second_time)
                     ).strftime("%Y/%m/%d %H:%M:%S")
            # 填补空缺
            str_time = str(time1) + '.000.'
            write_sheet.cell(row=i+1, column=1).value = str_time
            # 查看下一个
            i += 1
            second_time += 1
            now_value = read_sheet.row_values(i)
            now_time = now_value[0]
        else:
            i += 1
    workbook.save('./new_f2.xlsx')  # 保存新的excel,最终结果


mid_file = compared_data('./f1.xlsx', './f2.xlsx')
add_time(mid_file)
